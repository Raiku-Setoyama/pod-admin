"""Order list CSV/Excel generation utilities for manufacturer ordering."""

import csv
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Product type to Japanese name mapping
PRODUCT_TYPE_NAMES: dict[str, str] = {
    "acrylic_keychain": "アクリルキーホルダー",
    "acrylic_stand": "アクリルフィギュア",
    "sticker": "ステッカー",
    "tote_bag": "トートバッグ",
    "mug": "マグカップ",
    "tshirt": "Tシャツ",
}


def get_product_type_name(product_type: str) -> str:
    """Get Japanese name for product type."""
    return PRODUCT_TYPE_NAMES.get(product_type, product_type)


def format_product_detail(
    product_type: str,
    size: str | None = None,
    position: str | None = None,
    color: str | None = None,
) -> str:
    """Build the product detail string: {商品種類} - {サイズ} - {位置} - {色}.

    Elements without a value (size / position / color が null/空) are skipped so
    no extra separators appear. The product type name is resolved via
    get_product_type_name (e.g. acrylic_stand -> アクリルフィギュア).

    Examples:
        Tシャツ - L - 正面 - 白
        ステッカー - 100x100mm - ホワイト  (position 無し)
    """
    detail_parts = [get_product_type_name(product_type)]
    if size:
        detail_parts.append(size)
    if position:
        detail_parts.append(position)
    if color:
        detail_parts.append(color)
    return " - ".join(detail_parts)


class OrderListGenerator:
    """CSV generator for manufacturer order lists (発注リスト)."""

    def generate_order_list_csv(
        self,
        items: list[dict],
        product_type: str | None = None,
    ) -> bytes:
        """Generate order list CSV for a manufacturer.

        CSV Format:
        - 注文日: MM月DD日 形式
        - 注文番号: Order.order_number
        - 製品番号: OrderItem.uid
        - 商品名: OrderItem.product_name
        - 商品種類: {商品種類} - {サイズ} - {位置} - {色} (ex. Tシャツ - XL - 正面 - 白)
        - 原価: Product.cost
        - 個数: OrderItem.quantity
        - シール用テキスト情報: {商品名}【個数】{個数}個_{注文番号}_{製品番号}_{MMDD}

        Args:
            items: List of dicts containing order item data with keys:
                - ordered_date: datetime
                - order_number: str
                - uid: str
                - product_name: str
                - product_type: str (optional, used if product_type arg is None)
                - quantity: int
                - size: str | None
                - position: str | None
                - color: str | None
                - cost: int
            product_type: Product type code (e.g., "tshirt"). If None, uses
                          each item's product_type field.

        Returns:
            CSV content as bytes (UTF-8 with BOM for Excel compatibility).
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Column headers
        writer.writerow([
            "注文日",
            "注文番号",
            "製品番号",
            "商品名",
            "商品種類",
            "原価",
            "個数",
            "シール用テキスト情報",
        ])

        # Data rows
        for item in items:
            ordered_date: datetime | None = item.get("ordered_date")
            order_number = item.get("order_number", "")
            uid = item.get("uid", "")
            product_name = item.get("product_name", "")
            quantity = item.get("quantity", 1)
            size = item.get("size", "")
            position = item.get("position", "")
            color = item.get("color", "")
            cost = item.get("cost", 0)

            # Get product type (use argument if provided, else from item)
            item_product_type = product_type or item.get("product_type", "")

            # Format date as "MM月DD日"
            formatted_date = ordered_date.strftime("%m月%d日") if ordered_date else ""
            mmdd = ordered_date.strftime("%m%d") if ordered_date else ""

            # Build product detail: {商品種類} - {サイズ} - {位置} - {色}
            product_detail = format_product_detail(
                item_product_type, size, position, color
            )

            # Build seal text: {商品種類}【個数】{個数}個_{注文番号}_{製品番号}_{MMDD}
            seal_text = f"{product_detail}【個数】{quantity}個_{order_number}_{uid}_{mmdd}"

            writer.writerow([
                formatted_date,
                order_number,
                uid,
                product_name,
                product_detail,
                cost,
                quantity,
                seal_text,
            ])

        return output.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility

    def generate_order_list_xlsx(
        self,
        items: list[dict],
        product_type: str | None = None,
    ) -> bytes:
        """Generate order list Excel file for a manufacturer.

        Excel Format (6 columns, excludes 商品名 and 原価):
        - 注文日: M月D日 形式（ゼロ詰めなし、文字列形式）
        - 注文番号: Order.order_number
        - 製品番号: OrderItem.uid
        - 商品種類: {商品種類} - {サイズ} - {位置} - {色}
        - 個数: OrderItem.quantity
        - シール用テキスト情報: {商品名}【個数】{個数}個_{注文番号}_{製品番号}_{MMDD}

        Args:
            items: List of dicts containing order item data
            product_type: Product type code (e.g., "tshirt")

        Returns:
            Excel file content as bytes (.xlsx format).
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "発注リスト"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Column headers (6 columns, excluding 商品名 and 原価)
        headers = [
            "注文日",
            "注文番号",
            "製品番号",
            "商品種類",
            "個数",
            "シール用テキスト情報",
        ]
        column_widths = [12, 20, 15, 35, 8, 60]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, item in enumerate(items, 2):
            ordered_date: datetime | None = item.get("ordered_date")
            order_number = item.get("order_number", "")
            uid = item.get("uid", "")
            quantity = item.get("quantity", 1)
            size = item.get("size", "")
            position = item.get("position", "")
            color = item.get("color", "")

            # Get product type (use argument if provided, else from item)
            item_product_type = product_type or item.get("product_type", "")

            # Format date as "M月D日" (no zero-padding)
            if ordered_date:
                formatted_date = f"{ordered_date.month}月{ordered_date.day}日"
                mmdd = ordered_date.strftime("%m%d")
            else:
                formatted_date = ""
                mmdd = ""

            # Build product detail: {商品種類} - {サイズ} - {位置} - {色}
            product_detail = format_product_detail(
                item_product_type, size, position, color
            )

            # Build seal text: {商品種類}【個数】{個数}個_{注文番号}_{製品番号}_{MMDD}
            seal_text = f"{product_detail}【個数】{quantity}個_{order_number}_{uid}_{mmdd}"

            # Write row data
            # Column 1: 注文日 (as text string)
            date_cell = ws.cell(row=row_idx, column=1, value=formatted_date)
            date_cell.border = thin_border
            date_cell.number_format = "@"  # Text format

            # Column 2: 注文番号
            ws.cell(row=row_idx, column=2, value=order_number).border = thin_border

            # Column 3: 製品番号
            ws.cell(row=row_idx, column=3, value=uid).border = thin_border

            # Column 4: 商品種類
            ws.cell(row=row_idx, column=4, value=product_detail).border = thin_border

            # Column 5: 個数
            ws.cell(row=row_idx, column=5, value=quantity).border = thin_border

            # Column 6: シール用テキスト情報
            ws.cell(row=row_idx, column=6, value=seal_text).border = thin_border

        # Adjust column widths
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

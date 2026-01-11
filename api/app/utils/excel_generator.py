"""Excel generation utilities."""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelGenerator:
    """Excel file generator for purchase orders."""

    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def generate_purchase_order_excel(
        self,
        manufacturer_name: str,
        order_date: datetime,
        expected_delivery: datetime,
        items: list[dict],
    ) -> bytes:
        """Generate purchase order Excel file."""
        wb = Workbook()
        ws = wb.active
        ws.title = "発注書"

        # Header info
        ws["A1"] = "発注書"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:F1")

        ws["A3"] = "発注先:"
        ws["B3"] = manufacturer_name
        ws["A4"] = "発注日:"
        ws["B4"] = order_date.strftime("%Y/%m/%d")
        ws["A5"] = "納品予定日:"
        ws["B5"] = expected_delivery.strftime("%Y/%m/%d")

        # Table header
        headers = ["No.", "注文番号", "商品名", "サイズ", "数量", "単価", "小計"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal="center")

        # Table data
        total = 0
        for idx, item in enumerate(items, 1):
            row = idx + 7
            subtotal = item.get("price", 0) * item.get("quantity", 1)
            total += subtotal

            ws.cell(row=row, column=1, value=idx).border = self.thin_border
            ws.cell(row=row, column=2, value=item.get("order_number", "")).border = self.thin_border
            ws.cell(row=row, column=3, value=item.get("product_name", "")).border = self.thin_border
            ws.cell(row=row, column=4, value=item.get("size", "")).border = self.thin_border
            ws.cell(row=row, column=5, value=item.get("quantity", 1)).border = self.thin_border
            ws.cell(row=row, column=6, value=item.get("price", 0)).border = self.thin_border
            ws.cell(row=row, column=7, value=subtotal).border = self.thin_border

        # Total row
        total_row = len(items) + 8
        ws.cell(row=total_row, column=6, value="合計").font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=total).font = Font(bold=True)

        # Adjust column widths
        column_widths = [6, 15, 30, 12, 8, 10, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

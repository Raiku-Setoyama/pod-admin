"""Tests for order list generator."""

import io
from datetime import datetime
from typing import Any

import pytest
from openpyxl import load_workbook

from app.utils.order_list_generator import OrderListGenerator, get_product_type_name


class TestOrderListGenerator:
    """Tests for OrderListGenerator class."""

    @pytest.fixture
    def generator(self) -> Any:
        """Create generator instance."""
        return OrderListGenerator()

    def test_generate_order_list_csv_with_items(self, generator: Any) -> None:
        """Test CSV generation with items."""
        items = [
            {
                "ordered_date": datetime(2024, 12, 1, 10, 30, 0),
                "order_number": "2623928",
                "uid": "3765370",
                "product_name": "Tシャツ - XL - 正面 - 白",
                "product_type": "tshirt",
                "quantity": 1,
                "size": "XL",
                "position": "正面",
                "color": "白",
                "cost": 1500,
            },
            {
                "ordered_date": datetime(2024, 12, 1, 11, 0, 0),
                "order_number": "2624714",
                "uid": "3766653",
                "product_name": "Tシャツ - M - 正面 - 白",
                "product_type": "tshirt",
                "quantity": 2,
                "size": "M",
                "position": "正面",
                "color": "白",
                "cost": 1500,
            },
        ]

        result = generator.generate_order_list_csv(items)

        # Should be bytes
        assert isinstance(result, bytes)

        # Decode and check content
        content = result.decode("utf-8-sig")
        lines = [line.strip() for line in content.strip().split("\n")]

        # Check header
        assert lines[0] == "注文日,注文番号,製品番号,商品名,商品種類,原価,個数,シール用テキスト情報"

        # Check first data row
        assert "12月01日" in lines[1]
        assert "2623928" in lines[1]
        assert "3765370" in lines[1]
        assert "Tシャツ - XL - 正面 - 白" in lines[1]  # product_name
        assert "Tシャツ - XL - 正面 - 白" in lines[1]  # product_detail
        assert "1500" in lines[1]  # cost
        assert "【個数】1個" in lines[1]
        assert "_1201" in lines[1]  # MMDD format (12月01日)

        # Check second data row
        assert "12月01日" in lines[2]
        assert "2624714" in lines[2]
        assert "【個数】2個" in lines[2]

    def test_generate_order_list_csv_empty_items(self, generator: Any) -> None:
        """Test CSV generation with no items."""
        result = generator.generate_order_list_csv([])

        content = result.decode("utf-8-sig")
        lines = [line.strip() for line in content.strip().split("\n")]

        # Should only have header
        assert len(lines) == 1
        assert lines[0] == "注文日,注文番号,製品番号,商品名,商品種類,原価,個数,シール用テキスト情報"

    def test_generate_order_list_csv_seal_text_format(self, generator: Any) -> None:
        """Test seal text format is correct with MMDD and product_detail."""
        items = [
            {
                "ordered_date": datetime(2024, 12, 15, 10, 0, 0),
                "order_number": "ORD123",
                "uid": "PROD456",
                "product_name": "Test Product",
                "product_type": "tshirt",
                "quantity": 5,
                "size": "M",
                "position": "正面",
                "color": "白",
                "cost": 1000,
            },
        ]

        result = generator.generate_order_list_csv(items)
        content = result.decode("utf-8-sig")

        # Check seal text format: {商品種類}【個数】{個数}個_{注文番号}_{製品番号}_{MMDD}
        # 商品種類は product_detail (Tシャツ - M - 正面 - 白) を使用
        expected_seal = "Tシャツ - M - 正面 - 白【個数】5個_ORD123_PROD456_1215"
        assert expected_seal in content

    def test_generate_order_list_csv_date_format(self, generator: Any) -> None:
        """Test date format is MM月DD日 and seal text uses MMDD."""
        items = [
            {
                "ordered_date": datetime(2024, 1, 5, 10, 0, 0),
                "order_number": "ORD001",
                "uid": "PROD001",
                "product_name": "Product",
                "product_type": "tshirt",
                "quantity": 1,
                "size": "M",
                "position": "正面",
                "color": "白",
                "cost": 1000,
            },
        ]

        result = generator.generate_order_list_csv(items)
        content = result.decode("utf-8-sig")

        # Check date format with zero padding
        assert "01月05日" in content
        # Check seal text uses MMDD format (1月5日 -> 0105)
        assert "_0105" in content

    def test_generate_order_list_csv_utf8_bom(self, generator: Any) -> None:
        """Test CSV has UTF-8 BOM for Excel compatibility."""
        items = [
            {
                "ordered_date": datetime(2024, 12, 1, 10, 0, 0),
                "order_number": "ORD001",
                "uid": "PROD001",
                "product_name": "日本語テスト",
                "product_type": "tshirt",
                "quantity": 1,
                "size": "M",
                "position": "正面",
                "color": "白",
                "cost": 1000,
            },
        ]

        result = generator.generate_order_list_csv(items)

        # UTF-8 BOM is 0xEF, 0xBB, 0xBF
        assert result[:3] == b"\xef\xbb\xbf"

    def test_generate_order_list_csv_missing_optional_fields(self, generator: Any) -> None:
        """Test CSV generation with missing optional fields."""
        items = [
            {
                "ordered_date": None,
                "order_number": "",
                "uid": "",
                "product_name": "",
                "product_type": "",
                "quantity": 1,
                "size": None,
                "position": None,
                "color": None,
                "cost": 0,
            },
        ]

        result = generator.generate_order_list_csv(items)
        content = result.decode("utf-8-sig")

        # Should still generate CSV without errors
        lines = [line.strip() for line in content.strip().split("\n")]
        assert len(lines) == 2  # Header + 1 data row


class TestSealTextWithProductDetail:
    """AC-001: シール用テキスト情報が商品種類を使用し、MMDD形式で生成されるテスト"""

    @pytest.fixture
    def generator(self) -> Any:
        """Create generator instance."""
        return OrderListGenerator()

    def test_seal_text_uses_product_detail_and_mmdd_csv(self, generator: Any) -> None:
        """AC-001: シール用テキスト情報が商品種類を使用し、MMDD形式で生成される (CSV)

        given: 注文日が2026年2月15日、商品種類が「アクリルキーホルダー - 50x50 (mm) - アクリル」、
               個数が3、注文番号が「ORD-001」、製品番号が「UID-123」
        when: シール用テキスト情報を生成する
        then: 「アクリルキーホルダー - 50x50 (mm) - アクリル【個数】3個_ORD-001_UID-123_0215」が出力される
        """
        items = [
            {
                "ordered_date": datetime(2026, 2, 15, 10, 0, 0),
                "order_number": "ORD-001",
                "uid": "UID-123",
                "product_name": "アクリルキーホルダー",
                "product_type": "acrylic_keychain",
                "quantity": 3,
                "size": "50x50 (mm)",
                "position": None,
                "color": "アクリル",
                "cost": 1000,
            },
        ]

        result = generator.generate_order_list_csv(items)
        content = result.decode("utf-8-sig")

        # 期待されるシール用テキスト情報
        expected_seal = "アクリルキーホルダー - 50x50 (mm) - アクリル【個数】3個_ORD-001_UID-123_0215"
        assert expected_seal in content

    def test_seal_text_uses_product_detail_and_mmdd_xlsx(self, generator: Any) -> None:
        """AC-001: シール用テキスト情報が商品種類を使用し、MMDD形式で生成される (XLSX)

        given: 注文日が2026年2月15日、商品種類が「アクリルキーホルダー - 50x50 (mm) - アクリル」、
               個数が3、注文番号が「ORD-001」、製品番号が「UID-123」
        when: シール用テキスト情報を生成する
        then: 「アクリルキーホルダー - 50x50 (mm) - アクリル【個数】3個_ORD-001_UID-123_0215」が出力される
        """
        items = [
            {
                "ordered_date": datetime(2026, 2, 15, 10, 0, 0),
                "order_number": "ORD-001",
                "uid": "UID-123",
                "product_name": "アクリルキーホルダー",
                "product_type": "acrylic_keychain",
                "quantity": 3,
                "size": "50x50 (mm)",
                "position": None,
                "color": "アクリル",
                "cost": 1000,
            },
        ]

        result = generator.generate_order_list_xlsx(items)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active

        # 期待されるシール用テキスト情報
        expected_seal = "アクリルキーホルダー - 50x50 (mm) - アクリル【個数】3個_ORD-001_UID-123_0215"
        assert ws.cell(row=2, column=6).value == expected_seal

    def test_mmdd_format_december(self, generator: Any) -> None:
        """AC-003: MMDDは月の2桁と日の2桁で構成される（12月のケース）

        given: 注文日が2026年12月1日
        when: MMDD部分を生成する
        then: 「1201」が出力される
        """
        items = [
            {
                "ordered_date": datetime(2026, 12, 1, 10, 0, 0),
                "order_number": "ORD-001",
                "uid": "UID-123",
                "product_name": "テスト商品",
                "product_type": "tshirt",
                "quantity": 1,
                "size": "M",
                "position": "正面",
                "color": "白",
                "cost": 1000,
            },
        ]

        result = generator.generate_order_list_xlsx(items)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active

        seal_text = ws.cell(row=2, column=6).value
        assert "_1201" in seal_text

    def test_mmdd_format_january_with_padding(self, generator: Any) -> None:
        """AC-004: 1月の場合でもMMDDが正しく2桁で出力される

        given: 注文日が2027年1月5日
        when: MMDD部分を生成する
        then: 「0105」が出力される
        """
        items = [
            {
                "ordered_date": datetime(2027, 1, 5, 10, 0, 0),
                "order_number": "ORD-001",
                "uid": "UID-123",
                "product_name": "テスト商品",
                "product_type": "tshirt",
                "quantity": 1,
                "size": "M",
                "position": "正面",
                "color": "白",
                "cost": 1000,
            },
        ]

        result = generator.generate_order_list_xlsx(items)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active

        seal_text = ws.cell(row=2, column=6).value
        assert "_0105" in seal_text

    def test_seal_text_with_tshirt_full_product_detail(self, generator: Any) -> None:
        """Tシャツのシール用テキスト情報が正しく生成される

        given: Tシャツの注文
        when: シール用テキスト情報を生成する
        then: 商品種類「Tシャツ - M - 正面 - 白」が使用される
        """
        items = [
            {
                "ordered_date": datetime(2026, 2, 15, 10, 0, 0),
                "order_number": "ORD-002",
                "uid": "UID-456",
                "product_name": "Tシャツ",
                "product_type": "tshirt",
                "quantity": 2,
                "size": "M",
                "position": "正面",
                "color": "白",
                "cost": 1500,
            },
        ]

        result = generator.generate_order_list_xlsx(items)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active

        expected_seal = "Tシャツ - M - 正面 - 白【個数】2個_ORD-002_UID-456_0215"
        assert ws.cell(row=2, column=6).value == expected_seal


class TestGetProductTypeName:
    """Tests for get_product_type_name function."""

    def test_get_product_type_name_tshirt(self) -> None:
        """Test tshirt product type name."""
        assert get_product_type_name("tshirt") == "Tシャツ"

    def test_get_product_type_name_mug(self) -> None:
        """Test mug product type name."""
        assert get_product_type_name("mug") == "マグカップ"

    def test_get_product_type_name_acrylic_keychain(self) -> None:
        """Test acrylic_keychain product type name."""
        assert get_product_type_name("acrylic_keychain") == "アクリルキーホルダー"

    def test_get_product_type_name_acrylic_stand(self) -> None:
        """Test acrylic_stand product type name."""
        assert get_product_type_name("acrylic_stand") == "アクリルフィギュア"

    def test_get_product_type_name_sticker(self) -> None:
        """Test sticker product type name."""
        assert get_product_type_name("sticker") == "ステッカー"

    def test_get_product_type_name_tote_bag(self) -> None:
        """Test tote_bag product type name."""
        assert get_product_type_name("tote_bag") == "トートバッグ"

    def test_get_product_type_name_unknown(self) -> None:
        """Test unknown product type returns original value."""
        assert get_product_type_name("unknown_type") == "unknown_type"


class TestOrderListGeneratorXlsx:
    """Tests for OrderListGenerator XLSX generation."""

    @pytest.fixture
    def generator(self) -> Any:
        """Create generator instance."""
        return OrderListGenerator()

    def test_generate_order_list_xlsx_with_items(self, generator: Any) -> None:
        """Test XLSX generation with items."""
        items = [
            {
                "ordered_date": datetime(2024, 12, 1, 10, 30, 0),
                "order_number": "2623928",
                "uid": "3765370",
                "product_name": "Tシャツ - XL - 正面 - 白",
                "product_type": "tshirt",
                "quantity": 1,
                "size": "XL",
                "position": "正面",
                "color": "白",
                "cost": 1500,
            },
            {
                "ordered_date": datetime(2024, 12, 1, 11, 0, 0),
                "order_number": "2624714",
                "uid": "3766653",
                "product_name": "Tシャツ - M - 正面 - 白",
                "product_type": "tshirt",
                "quantity": 2,
                "size": "M",
                "position": "正面",
                "color": "白",
                "cost": 1500,
            },
        ]

        result = generator.generate_order_list_xlsx(items)

        # Should be bytes
        assert isinstance(result, bytes)

        # Load workbook and check content
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active

        # Check sheet name
        assert ws.title == "発注リスト"

        # Check header (6 columns, no 商品名 or 原価)
        headers = [ws.cell(row=1, column=i).value for i in range(1, 7)]
        assert headers == [
            "注文日",
            "注文番号",
            "製品番号",
            "商品種類",
            "個数",
            "シール用テキスト情報",
        ]

        # Check first data row
        assert ws.cell(row=2, column=1).value == "12月1日"  # No zero-padding
        assert ws.cell(row=2, column=2).value == "2623928"
        assert ws.cell(row=2, column=3).value == "3765370"
        assert ws.cell(row=2, column=4).value == "Tシャツ - XL - 正面 - 白"
        assert ws.cell(row=2, column=5).value == 1
        assert "【個数】1個" in ws.cell(row=2, column=6).value

        # Check second data row
        assert ws.cell(row=3, column=1).value == "12月1日"
        assert ws.cell(row=3, column=2).value == "2624714"
        assert ws.cell(row=3, column=5).value == 2

    def test_generate_order_list_xlsx_empty_items(self, generator: Any) -> None:
        """Test XLSX generation with no items."""
        result = generator.generate_order_list_xlsx([])

        wb = load_workbook(io.BytesIO(result))
        ws = wb.active

        # Should only have header
        assert ws.cell(row=1, column=1).value == "注文日"
        assert ws.cell(row=2, column=1).value is None

    def test_generate_order_list_xlsx_date_format_no_zero_padding(self, generator: Any) -> None:
        """Test date format is M月D日 (no zero-padding)."""
        items = [
            {
                "ordered_date": datetime(2024, 1, 5, 10, 0, 0),
                "order_number": "ORD001",
                "uid": "PROD001",
                "product_name": "Product",
                "product_type": "tshirt",
                "quantity": 1,
                "size": "M",
                "position": "正面",
                "color": "白",
                "cost": 1000,
            },
        ]

        result = generator.generate_order_list_xlsx(items)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active

        # Check date format without zero padding
        assert ws.cell(row=2, column=1).value == "1月5日"

    def test_generate_order_list_xlsx_date_cell_is_text_format(self, generator: Any) -> None:
        """Test date cell is formatted as text."""
        items = [
            {
                "ordered_date": datetime(2024, 12, 15, 10, 0, 0),
                "order_number": "ORD123",
                "uid": "PROD456",
                "product_name": "Test Product",
                "product_type": "tshirt",
                "quantity": 5,
                "size": "M",
                "position": "正面",
                "color": "白",
                "cost": 1000,
            },
        ]

        result = generator.generate_order_list_xlsx(items)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active

        # Check date cell number format is text (@)
        assert ws.cell(row=2, column=1).number_format == "@"

    def test_generate_order_list_xlsx_seal_text_format(self, generator: Any) -> None:
        """Test seal text format is correct with MMDD and product_detail."""
        items = [
            {
                "ordered_date": datetime(2024, 12, 15, 10, 0, 0),
                "order_number": "ORD123",
                "uid": "PROD456",
                "product_name": "Test Product",
                "product_type": "tshirt",
                "quantity": 5,
                "size": "M",
                "position": "正面",
                "color": "白",
                "cost": 1000,
            },
        ]

        result = generator.generate_order_list_xlsx(items)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active

        # Check seal text format: {商品種類}【個数】{個数}個_{注文番号}_{製品番号}_{MMDD}
        # 商品種類は product_detail (Tシャツ - M - 正面 - 白) を使用
        expected_seal = "Tシャツ - M - 正面 - 白【個数】5個_ORD123_PROD456_1215"
        assert ws.cell(row=2, column=6).value == expected_seal

    def test_generate_order_list_xlsx_excludes_product_name_and_cost(self, generator: Any) -> None:
        """Test XLSX excludes 商品名 and 原価 columns."""
        items = [
            {
                "ordered_date": datetime(2024, 12, 1, 10, 0, 0),
                "order_number": "ORD001",
                "uid": "PROD001",
                "product_name": "商品名テスト",
                "product_type": "tshirt",
                "quantity": 1,
                "size": "M",
                "position": "正面",
                "color": "白",
                "cost": 9999,  # This should NOT appear in XLSX
            },
        ]

        result = generator.generate_order_list_xlsx(items)
        wb = load_workbook(io.BytesIO(result))
        ws = wb.active

        # Check that we have exactly 6 columns
        headers = [ws.cell(row=1, column=i).value for i in range(1, 10)]
        non_none_headers = [h for h in headers if h is not None]
        assert len(non_none_headers) == 6

        # Verify 商品名 is NOT a column header
        assert "商品名" not in headers

        # Verify 原価 is NOT a column header
        assert "原価" not in headers

        # Verify cost value (9999) does NOT appear in the row
        row_values = [ws.cell(row=2, column=i).value for i in range(1, 7)]
        assert 9999 not in row_values

#!/usr/bin/env python3
"""export-xlsx.py — 要件・検討事項・決定事項・不具合を 1 つの Excel（4 タブ）で出力する。

使い方:
    python3 scripts/export-xlsx.py              # dist/プロジェクト管理表.xlsx を出す
    python3 scripts/export-xlsx.py --internal   # 社内用（内部項目も含める）
    python3 scripts/export-xlsx.py --out dist   # 出力先ディレクトリを指定

列の定義は _exports.py に集約している（CSV 版の export-csv.py と共通）。
顧客向け（既定）ではステータスを日本語に訳し、内部管理用の項目を落とす。

依存: pip install openpyxl
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from _exports import BUILDERS

DOCS = Path("docs")
OUT = Path("dist")
FILENAME = "プロジェクト管理表.xlsx"

# 列幅の自動調整の上限。長い本文セルでシートが横に間延びするのを防ぐ。
MAX_COL_WIDTH = 60


def build_workbook(client: bool):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)  # 既定の空シートを消す

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    wrap = Alignment(vertical="top", wrap_text=True)

    counts: dict[str, int] = {}
    for _label, sheet_name, builder in BUILDERS.values():
        header, rows = builder(client=client)
        ws = wb.create_sheet(title=sheet_name)
        ws.append(header)
        for row in rows:
            ws.append(row)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        # 見出し行を固定して、スクロールしても列名が見えるようにする
        ws.freeze_panes = "A2"
        # 顧客が領域・状況で絞り込めるようにする
        ws.auto_filter.ref = ws.dimensions

        for idx, _name in enumerate(header, start=1):
            column = [str(ws.cell(row=r, column=idx).value or "")
                      for r in range(1, ws.max_row + 1)]
            width = min(max((len(v) for v in column), default=8) + 2, MAX_COL_WIDTH)
            ws.column_dimensions[get_column_letter(idx)].width = width
        for row_cells in ws.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = wrap

        counts[sheet_name] = len(rows)
    return wb, counts


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--internal", action="store_true", help="内部項目も含めて出力する")
    parser.add_argument("--out", default=str(OUT))
    args = parser.parse_args()

    if not DOCS.exists():
        print("docs/ が見つかりません。リポジトリのルートで実行してください。", file=sys.stderr)
        return 1

    try:
        import openpyxl  # noqa: F401
    except ModuleNotFoundError:
        print("openpyxl が必要です。`pip install openpyxl` を実行してください。", file=sys.stderr)
        return 1

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)
    stem, suffix = FILENAME.rsplit(".", 1)
    name = f"{stem}{'-internal' if args.internal else ''}.{suffix}"
    target = out_dir / name

    wb, counts = build_workbook(client=not args.internal)
    wb.save(target)

    summary = " / ".join(f"{sheet} {n} 件" for sheet, n in counts.items())
    print(f"{target} を出力しました（{summary}）")
    return 0


if __name__ == "__main__":
    sys.exit(main())
